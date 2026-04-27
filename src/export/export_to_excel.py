import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


from src.db.db import get_connection
from src.normalization.normalizer import normalize_extraction
from src.business_rules.calculations import (
    compute_cov_t,
    compute_max_quantity_t,
    default_cov_percent,
    format_decimal,
    parse_number,
)


TEMPLATE_PATH = "data/templates/_SEVESO e COV_14012026.xlsx"
OUTPUT_PATH = "data/output/seveso_cov_export.xlsx"
SHEET_NAME = "MP Qumicos Res"


HEADERS = [
    "Nº",
    "Modo potencial de emissão ",
    "Referencia",
    "Designação da substância / mistura / resíduo / água residual",
    "Tipo de substância / Utilização",
    "Local de utilização / produção ",
    "Consumo anual / Produção anual (2025)",
    "Un",
    "t",
    "Capacidade de Armazenamento",
    "Un",
    "Capacidade de Armazenamento (t)",
    "COV g/l",
    "COV %",
    "Consumo COV (t)",
    "Constituintes da mistura",
    "% na mistura",
    "N.º CAS",
    "N.º CE",
    "Classificação da substância / mistura / resíduo / água residual  ",
    "Densidade (g/cm3)",
    "Quantidade máxima armazenada (t)",
    "SEVESO",
    "N.versão",
    "Data",
    "Fornecedor",
    "Lingua",
]

HAZARD_CLASS_ABBREVIATIONS = {
    "H225": "Flam. Liq. 2",
    "H226": "Flam. Liq. 3",
    "H304": "Asp. Tox. 1",
    "H312": "Acute Tox. 4",
    "H315": "Skin Irrit. 2",
    "H317": "Skin Sens. 1",
    "H318": "Eye Dam. 1",
    "H319": "Eye Irrit. 2",
    "H332": "Acute Tox. 4",
    "H335": "STOT SE 3",
    "H336": "STOT SE 3",
    "H361d": "Repr. 2",
    "H373": "STOT RE 2",
    "H412": "Aquatic Chronic 3",
}


def clean_excel_placeholder(value):
    """Remove placeholder values from generated Excel data cells."""
    if value is None:
        return ""

    value = str(value).strip()
    if not value:
        return ""

    placeholders = {
        "não disponível",
        "nao disponivel",
        "não aplicável",
        "nao aplicavel",
        "no disponible",
        "not available",
        "not applicable",
        "none",
        "null",
        "n/a",
    }

    if value.lower() in placeholders:
        return ""

    return value


def format_substance_for_excel(substance):
    """Format one substance name for the Excel constituents column.

    CAS numbers and percentages are exported in their own columns, so this column
    should contain only the constituent name.
    """
    name = clean_excel_placeholder(substance.get("name"))

    if not name:
        return ""

    return f"• {name}"


def format_substances_for_excel(substances):
    """Format all substances as a readable multiline Excel cell."""
    formatted = []

    for substance in substances:
        value = format_substance_for_excel(substance)
        if value:
            formatted.append(value)

    return "\n".join(formatted)


def ensure_export_columns(conn):
    """Ensure optional export columns exist even when export runs standalone."""
    cur = conn.cursor()
    cur.execute("PRAGMA table_info(extractions)")
    existing_columns = {row[1] for row in cur.fetchall()}

    required_columns = {
        "annual_consumption_kg": "TEXT",
        "storage_capacity_l": "TEXT",
    }

    for column_name, column_type in required_columns.items():
        if column_name not in existing_columns:
            cur.execute(f"ALTER TABLE extractions ADD COLUMN {column_name} {column_type}")

    conn.commit()

def fetch_extractions():
    conn = get_connection()
    ensure_export_columns(conn)
    cur = conn.cursor()

    cur.execute("""
        SELECT
            file_name,
            extracted_json,
            annual_consumption_kg,
            storage_capacity_l
        FROM extractions
        WHERE status = 'approved'
        ORDER BY id ASC
    """)

    rows = cur.fetchall()
    conn.close()
    return rows

# Helper to safely get from row (sqlite3.Row or dict)
def row_get(row, key, default=""):
    """Safely get a value from sqlite3.Row or dict-like rows."""
    try:
        return row[key]
    except (KeyError, IndexError, TypeError):
        return default


def join_values(items, key):
    values = []

    fallback_keys = {
        "ce_number": ["ce_number", "ec_number", "einecs_number", "ec", "ce"],
        "cas_number": ["cas_number", "cas", "cas_no"],
    }

    candidate_keys = fallback_keys.get(key, [key])

    for item in items:
        value = ""
        for candidate_key in candidate_keys:
            value = clean_excel_placeholder(item.get(candidate_key))
            if value:
                break

        if value:
            values.append(str(value))

    return "\n".join(values)

def restore_ce_numbers(normalized_substances, raw_substances):
    """Restore CE numbers that may be dropped during normalization.

    The raw extraction JSON can contain ce_number, while normalize_extraction may
    return substances without that key. Match first by CAS, then by lowercase name.
    """
    ce_by_cas = {}
    ce_by_name = {}

    for substance in raw_substances or []:
        ce_number = clean_excel_placeholder(substance.get("ce_number"))
        if not ce_number:
            continue

        cas_number = clean_excel_placeholder(substance.get("cas_number"))
        name = clean_excel_placeholder(substance.get("name")).lower()

        if cas_number:
            ce_by_cas[cas_number] = ce_number
        if name:
            ce_by_name[name] = ce_number

    for substance in normalized_substances or []:
        if clean_excel_placeholder(substance.get("ce_number")):
            continue

        cas_number = clean_excel_placeholder(substance.get("cas_number"))
        name = clean_excel_placeholder(substance.get("name")).lower()

        if cas_number and cas_number in ce_by_cas:
            substance["ce_number"] = ce_by_cas[cas_number]
        elif name and name in ce_by_name:
            substance["ce_number"] = ce_by_name[name]

    return normalized_substances


def format_hazard_classifications(hazards):
    """Format hazards for Excel using only H/EUH hazard statement codes."""
    values = []
    seen = set()

    for hazard in hazards:
        code = clean_excel_placeholder(hazard.get("code"))
        description = clean_excel_placeholder(hazard.get("description"))

        if not code:
            continue

        if not re.fullmatch(r"(?:H\d{3}[a-z]?|EUH\d{3}[a-z]?)", code):
            continue

        if code in seen:
            continue

        seen.add(code)

        if description:
            values.append(f"{code}: {description}")
        else:
            values.append(code)

    return "\n".join(values)


def compute_annual_consumption_t(consumption_kg):
    """Convert annual consumption from kg to tonnes for Excel."""
    value = parse_number(consumption_kg)

    if value is None:
        return ""

    return format_decimal(value / 1000, digits=3)


# Helper function to infer default COV percentage for Excel export.
def infer_default_cov_percent(cov_g_l, hazards):
    """Infer default COV percentage for Excel export.

    If an explicit COV g/L value exists, the project business rule sets COV % to
    100%. If COV g/L is missing but the mixture is classified as flammable
    liquid/vapour (H226/H225), use the same 100% default for solvent/thinner-like
    products.
    """
    explicit_percent = default_cov_percent(cov_g_l)

    if explicit_percent:
        return explicit_percent

    for hazard in hazards or []:
        code = clean_excel_placeholder(hazard.get("code"))
        if code in {"H225", "H226"}:
            return "100,00"

    return ""


def build_excel_rows(extractions):
    rows = []

    for index, row in enumerate(extractions, start=1):
        raw_data = json.loads(row["extracted_json"])
        data = normalize_extraction(raw_data)
        metadata = data.get("document_metadata", {}) or {}
        print(
            f"{row['file_name']} -> {len(data.get('substances', []))} substances | "
            f"validation: {data.get('validation_status', 'REVIEW')}"
        )

        substances = restore_ce_numbers(
            data.get("substances", []),
            raw_data.get("substances", []),
        )
        hazards = data.get("hazards", [])
        transport = data.get("transport", {}) or {}
        seveso = data.get("seveso", {}) or {}
        physical = data.get("physical_properties", {}) or {}

        operational = data.get("operational_data", {}) or {}

        annual_consumption_kg = clean_excel_placeholder(
            operational.get("annual_consumption_kg") or row_get(row, "annual_consumption_kg")
        )
        storage_capacity_l = clean_excel_placeholder(
            operational.get("storage_capacity_l") or row_get(row, "storage_capacity_l")
        )

        density = clean_excel_placeholder(physical.get("density", ""))
        cov_g_l = clean_excel_placeholder(physical.get("voc", ""))

        annual_consumption_t = compute_annual_consumption_t(annual_consumption_kg)
        storage_capacity_t = compute_max_quantity_t(storage_capacity_l, density)
        cov_percent = infer_default_cov_percent(cov_g_l, hazards)
        cov_t = compute_cov_t(annual_consumption_kg, cov_percent)
        max_quantity_t = storage_capacity_t

        hazard_codes = format_hazard_classifications(hazards)

        row_data = [
            index,
            "Uso, armazenamento e libertação",
            clean_excel_placeholder(metadata.get("product_code")),
            clean_excel_placeholder(metadata.get("product_name")) or row["file_name"].replace(".pdf", ""),
            "Matéria prima",
            "",
            annual_consumption_kg,
            "kg",
            annual_consumption_t,
            storage_capacity_l,
            "l",
            storage_capacity_t,
            cov_g_l,
            cov_percent,
            cov_t,
            format_substances_for_excel(substances),
            join_values(substances, "percentage"),
            join_values(substances, "cas_number"),
            join_values(substances, "ce_number"),
            hazard_codes,
            density,
            max_quantity_t,
            clean_excel_placeholder(seveso.get("category", "")),
            clean_excel_placeholder(metadata.get("version")),
            clean_excel_placeholder(metadata.get("revision_date")),
            clean_excel_placeholder(metadata.get("supplier")),
            clean_excel_placeholder(metadata.get("language")) or data.get("language") or "Português",
        ]

        if len(row_data) != len(HEADERS):
            print("DEBUG MISMATCH:", len(row_data), "vs", len(HEADERS))
        rows.append(row_data)

    return rows


def clear_old_rows(ws, start_row=3):
    if ws.max_row >= start_row:
        ws.delete_rows(start_row, ws.max_row - start_row + 1)


def copy_row_style(ws, source_row, target_row, max_col):
    for col in range(1, max_col + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def export_to_excel():
    Path("data/output").mkdir(parents=True, exist_ok=True)

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    template_style_row = 3

    style_values = [
        copy(ws.cell(row=template_style_row, column=col)._style)
        for col in range(1, len(HEADERS) + 1)
    ]

    clear_old_rows(ws, start_row=3)

    rows = build_excel_rows(fetch_extractions())
    if not rows:
        print("No approved extractions found. Excel was generated with no data rows.")

    for row_index, row_values in enumerate(rows, start=3):
        if len(row_values) != len(HEADERS):
            raise ValueError(
                f"Row {row_index} has {len(row_values)} values, but template expects {len(HEADERS)} columns."
            )

    for row_index, row_values in enumerate(rows, start=3):
        for col_index, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = value
            cell._style = copy(style_values[col_index - 1])
            cell.alignment = copy(cell.alignment.copy(wrap_text=True, vertical="top"))

    for row_index in range(3, 3 + len(rows)):
        ws.row_dimensions[row_index].height = 180

    ws.column_dimensions["P"].width = 85
    ws.column_dimensions["Q"].width = 18
    ws.column_dimensions["R"].width = 18
    ws.column_dimensions["T"].width = 30

    wb.save(OUTPUT_PATH)
    print(f"Excel exported to: {OUTPUT_PATH}")


if __name__ == "__main__":
    export_to_excel()